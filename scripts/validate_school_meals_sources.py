#!/usr/bin/env python3
"""Validate school-meals raw snapshots before transformation."""
from __future__ import annotations
import pathlib, sys
from openpyxl import load_workbook
ROOT=pathlib.Path(__file__).resolve().parents[1]
RAW=ROOT/'raw'/'school_meals'
REQ={'SML_TOTAL_CHILDREN','EDU_ENR_PRI','EDU_ENR_PRE','QLT_GDQS_MENU'}
found=set()
if not RAW.exists():
    print('[school raw QA FAIL] raw/school_meals does not exist', file=sys.stderr); sys.exit(1)
for path in RAW.glob('*.xlsx'):
    wb=load_workbook(path, read_only=True, data_only=True)
    ws=wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[0]]
    rows=ws.iter_rows(values_only=True)
    header=[str(x or '').strip() for x in next(rows)]
    idx={h:i for i,h in enumerate(header)}
    if 'Indicator Code' not in idx or 'ISO' not in idx or 'Year' not in idx or 'Value' not in idx:
        print(f'[school raw QA FAIL] missing standard columns: {path.name}', file=sys.stderr); sys.exit(1)
    codes=set(); valid=0
    for row in rows:
        code=str(row[idx['Indicator Code']] or '').strip()
        if code: codes.add(code)
        val=row[idx['Value']]
        if val not in (None,'—','-',''):
            valid+=1
    found |= codes
    print(f'[school raw QA] {path.name}: codes={sorted(codes)} valid_rows={valid}')
missing=REQ-found
if missing:
    print('[school raw QA FAIL] missing required indicators: '+', '.join(sorted(missing)), file=sys.stderr); sys.exit(1)
print('[school raw QA PASS] required school-meals snapshots present')
