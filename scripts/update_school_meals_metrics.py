#!/usr/bin/env python3
"""Hardened MARKET BASE school-meals transformer.

Purpose
-------
Generate stable Full196 school-meals JSON and rankings from official/portal XLSX or CSV exports.
This is intentionally defensive because WFP/SMC/GCNF school-meals data exports are less stable
than FAOSTAT bulk files.

Inputs
------
raw/school_meals/ may contain XLSX or CSV files with any of these shapes:
  - Data sheet columns: Indicator Code, ISO, Country, Year, Value, Unit, Source
  - Lowercase variants: indicator_code, iso3, country, year, value, unit, source
  - One indicator per file, where the indicator code is in the file name

Supported indicators
--------------------
  SML_TOTAL_CHILDREN  -> school_meals_total_children
  EDU_ENR_PRI         -> primary_enrollment_children
  EDU_ENR_PRE         -> pre_primary_enrollment_children
  QLT_GDQS_MENU       -> school_meal_menu_gdqs_score

Outputs
-------
  data/market_base_school_meals_data_CURRENT.json
  data/market_base_school_meals_data_v1.json / v94.json compatibility copies
  data/market_base_school_meals_rankings_CURRENT.json
  data/market_base_school_meals_rankings_v1.json compatibility copy
  data/market_base_school_meals_coverage_report_v103.json
  data/market_base_school_meals_source_manifest_v103.json

Safety policy
-------------
  - Full196 is always preserved.
  - Empty/dash values never overwrite valid existing values.
  - Latest available valid year is selected per ISO3 + indicator.
  - If two values share the latest year, the row with a non-empty source wins.
  - Existing data is preserved when no raw files are present.
  - Derived ratio is explicitly labelled as indicative, not a true coverage rate.
"""
from __future__ import annotations
import csv, json, pathlib, datetime, re, hashlib, sys
from collections import defaultdict
try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

ROOT = pathlib.Path(__file__).resolve().parents[1]
DATA = ROOT / 'data'
RAW = ROOT / 'raw' / 'school_meals'
CREATED_AT = datetime.date.today().isoformat()
VERSION = 'v103_school_meals_hardened'

INDICATORS = {
    'SML_TOTAL_CHILDREN': {
        'key':'school_meals_total_children', 'unit':'children', 'label_ja':'学校給食対象児童数',
        'label_en':'Children receiving school meals (total)'
    },
    'EDU_ENR_PRI': {
        'key':'primary_enrollment_children', 'unit':'children', 'label_ja':'小学校就学者数',
        'label_en':'Enrolment, primary'
    },
    'EDU_ENR_PRE': {
        'key':'pre_primary_enrollment_children', 'unit':'children', 'label_ja':'就学前教育在籍者数',
        'label_en':'Enrolment, pre-primary'
    },
    'QLT_GDQS_MENU': {
        'key':'school_meal_menu_gdqs_score', 'unit':'score', 'label_ja':'メニュー品質スコア',
        'label_en':'Global Diet Quality Score - menu'
    },
}
KEY_TO_CODE = {v['key']: k for k, v in INDICATORS.items()}
SCHOOL_METRICS = {v['key']:{'label':v['label_ja'],'unit':v['unit']} for v in INDICATORS.values()}
SCHOOL_METRICS['school_meal_children_per_primary_enrollment_pct'] = {'label':'学校給食対象児童数 / 小学校就学者数','unit':'%'}

DASHES = {'', '-', '—', '–', 'null', 'None', 'nan', 'NaN', 'N/A', 'n/a'}

def load_json(path):
    with open(path, encoding='utf-8') as f: return json.load(f)

def save_json(path, obj):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f: json.dump(obj, f, ensure_ascii=False, indent=2)

def norm_header(s):
    return re.sub(r'[^a-z0-9]+', '_', str(s or '').strip().lower()).strip('_')

def first(row, *names):
    for n in names:
        if n in row: return row.get(n)
    return None

def num(v):
    if v is None: return None
    if isinstance(v, (int, float)):
        return float(v)
    s=str(v).strip().replace(',', '')
    if s in DASHES: return None
    # tolerate values like "1 234" and strings with footnote spaces
    s=re.sub(r'\s+', '', s)
    try: return float(s)
    except Exception: return None

def clean_source(v):
    s=str(v or '').strip()
    return None if s in DASHES else s

def infer_indicator_from_name(path):
    upper=path.name.upper()
    for code in INDICATORS:
        if code in upper: return code
    return None

def file_hash(path):
    h=hashlib.sha256()
    with open(path,'rb') as f:
        for chunk in iter(lambda:f.read(1024*1024), b''):
            h.update(chunk)
    return h.hexdigest()

def iter_rows(path):
    if path.suffix.lower() == '.xlsx':
        if load_workbook is None:
            raise RuntimeError('openpyxl is required to read xlsx files')
        wb=load_workbook(path, read_only=True, data_only=True)
        ws=wb['Data'] if 'Data' in wb.sheetnames else wb[wb.sheetnames[0]]
        rows=ws.iter_rows(values_only=True)
        try:
            header_raw=next(rows)
        except StopIteration:
            return
        header=[norm_header(x) for x in header_raw]
        for values in rows:
            yield {header[i]: values[i] if i < len(values) else None for i in range(len(header))}
    else:
        with open(path, encoding='utf-8-sig', newline='') as f:
            reader=csv.DictReader(f)
            for raw in reader:
                yield {norm_header(k): v for k,v in raw.items()}

def read_metadata(path):
    meta=[]
    if path.suffix.lower() != '.xlsx' or load_workbook is None:
        return meta
    try:
        wb=load_workbook(path, read_only=True, data_only=True)
        if 'Metadata' not in wb.sheetnames: return meta
        ws=wb['Metadata']
        rows=ws.iter_rows(values_only=True)
        header=[norm_header(x) for x in next(rows)]
        for values in rows:
            row={header[i]: values[i] if i < len(values) else None for i in range(len(header))}
            if any(v is not None for v in row.values()): meta.append(row)
    except Exception as e:
        meta.append({'metadata_read_error': str(e)})
    return meta

def latest_values(files):
    out={}
    manifest=[]
    row_stats=[]
    for path in files:
        inferred=infer_indicator_from_name(path)
        m={'file':path.name,'bytes':path.stat().st_size,'sha256':file_hash(path),'metadata':read_metadata(path)}
        manifest.append(m)
        total=valid=accepted=0
        print(f'[school] source: {path}')
        for row in iter_rows(path):
            total += 1
            code=str(first(row,'indicator_code','indicator','code') or inferred or '').strip().upper()
            if code not in INDICATORS: continue
            iso=str(first(row,'iso','iso3','country_code','area_code_m49') or '').strip().upper()
            if len(iso) != 3: continue
            val=num(first(row,'value','obs_value','data_value'))
            if val is None: continue
            y=first(row,'year','time_period','time','period')
            try: year=int(float(str(y).strip()))
            except Exception: continue
            valid += 1
            key=INDICATORS[code]['key']
            unit=str(first(row,'unit','unit_of_measure') or INDICATORS[code]['unit']).strip() or INDICATORS[code]['unit']
            source=clean_source(first(row,'source','sources','source_detail'))
            country=clean_source(first(row,'country','area','country_name'))
            candidate={'value':val,'year':year,'unit':INDICATORS[code]['unit'],'display_unit':unit,'source':source or 'School Meals Coalition / WFP / GCNF export','indicator_code':code,'source_file':path.name}
            if country: candidate['source_country_name']=country
            prev=out.get((iso,key))
            use=False
            if prev is None: use=True
            elif year > (prev.get('year') or -9999): use=True
            elif year == (prev.get('year') or -9999) and source and not prev.get('source'):
                use=True
            if use:
                out[(iso,key)]=candidate
                accepted += 1
        row_stats.append({'file':path.name,'rows_total':total,'rows_valid_values':valid,'accepted_or_replaced':accepted})
    return out, manifest, row_stats

def build_rankings(records):
    rankings={'meta':{'dataset':'market_base_school_meals_rankings','version':VERSION,'created_at':CREATED_AT,'ranking_policy':'Generated from market_base_school_meals_data_CURRENT.json; missing values excluded; derived ratio is indicative only.'},'rankings':{}}
    for k,cfg in SCHOOL_METRICS.items():
        items=[]
        for r in records:
            obj=r.get(k)
            if not isinstance(obj, dict) or obj.get('value') is None: continue
            try: v=float(obj['value'])
            except Exception: continue
            items.append({'entity_id':r.get('entity_id'),'iso3':r.get('iso3'),'name_ja':r.get('name_ja'),'name_en':r.get('name_en'),'value':v,'year':obj.get('year'),'source':obj.get('source'),'unit':obj.get('unit') or cfg['unit']})
        items.sort(key=lambda x:x['value'], reverse=True)
        for i,it in enumerate(items,1): it['rank']=i
        rankings['rankings'][k]={'metric_name_ja':cfg['label'],'unit':cfg['unit'],'items':items}
    return rankings

def main():
    entities=load_json(DATA/'market_base_entities_basic_stats_full196_rc.json')['entities']
    existing_path=DATA/'market_base_school_meals_data_CURRENT.json'
    school=load_json(existing_path) if existing_path.exists() else {'meta':{}, 'records':[]}
    existing={r.get('entity_id'):r for r in school.get('records', [])}
    records=[]
    for e in entities:
        r=dict(existing.get(e['entity_id']) or {'entity_id':e['entity_id']})
        r.update({'iso3':e.get('iso3'), 'iso2':e.get('iso2'), 'name_ja':e.get('names',{}).get('ja') or r.get('name_ja'), 'name_en':e.get('names',{}).get('en') or r.get('name_en')})
        records.append(r)
    rec_by_iso={r.get('iso3'):r for r in records if r.get('iso3')}
    files=[]
    if RAW.exists():
        files=sorted([p for p in RAW.iterdir() if p.suffix.lower() in {'.xlsx','.csv'}])
    values={}; manifest=[]; row_stats=[]
    if files:
        values, manifest, row_stats = latest_values(files)
        unknown_iso=sorted({iso for iso,_ in values.keys() if iso not in rec_by_iso})
        if unknown_iso:
            print('[school] warning: raw source has ISO3 not in Full196:', ', '.join(unknown_iso[:20]))
        for (iso,key), obj in values.items():
            if iso in rec_by_iso:
                # Only overwrite with valid, newer/equal source values. Empty values never reach this point.
                prev=rec_by_iso[iso].get(key)
                if not isinstance(prev, dict) or obj.get('year',0) >= (prev.get('year') or 0):
                    rec_by_iso[iso][key]=obj
    else:
        print('[school] no raw/school_meals files found; keeping existing school-meals JSON and regenerating rankings')

    for r in records:
        total=(r.get('school_meals_total_children') or {}).get('value') if isinstance(r.get('school_meals_total_children'), dict) else None
        prim=(r.get('primary_enrollment_children') or {}).get('value') if isinstance(r.get('primary_enrollment_children'), dict) else None
        if total is not None and prim:
            try:
                if float(prim)>0:
                    r['school_meal_children_per_primary_enrollment_pct']={
                        'value':round(float(total)/float(prim)*100,2),
                        'year':max((r.get('school_meals_total_children') or {}).get('year') or 0,(r.get('primary_enrollment_children') or {}).get('year') or 0),
                        'unit':'%',
                        'source':'Derived from school-meals and education indicators',
                        'indicator_code':'SML_TOTAL_CHILDREN / EDU_ENR_PRI',
                        'note':'Indicative ratio only; denominator is primary enrollment and may not match exact program target population.'}
            except Exception: pass
        loaded=sum(1 for k in SCHOOL_METRICS if isinstance(r.get(k), dict) and r[k].get('value') is not None)
        r['school_meals_metrics_loaded_count']=loaded
        r['data_status']='loaded_full' if loaded>=4 else ('loaded_partial' if loaded else 'not_loaded')
    school['records']=records
    school.setdefault('meta',{}).update({
        'version':VERSION,
        'created_at':CREATED_AT,
        'source_policy':'Generated from official school-meals / education XLSX or CSV exports placed in raw/school_meals/. Included raw snapshots are used as stable fallback when portal bulk download is not available.',
        'ranking_policy':'School-meals rankings are generated from this JSON; missing values excluded; derived ratio is indicative only.',
        'full196_preserved':True,
        'raw_files_used':[m['file'] for m in manifest],
    })
    save_json(DATA/'market_base_school_meals_data_CURRENT.json', school)
    save_json(DATA/'market_base_school_meals_data_v1.json', school)
    save_json(DATA/'market_base_school_meals_data_v94.json', school)
    rankings = build_rankings(records)
    save_json(DATA/'market_base_school_meals_rankings_CURRENT.json', rankings)
    save_json(DATA/'market_base_school_meals_rankings_v1.json', rankings)
    coverage={'meta':{'version':VERSION,'created_at':CREATED_AT},'metric_coverage':{},'country_status_counts':{},'row_stats':row_stats}
    for k in SCHOOL_METRICS:
        coverage['metric_coverage'][k]=sum(1 for r in records if isinstance(r.get(k), dict) and r[k].get('value') is not None)
    for r in records:
        coverage['country_status_counts'][r['data_status']]=coverage['country_status_counts'].get(r['data_status'],0)+1
    coverage['low_coverage_countries']=[{'entity_id':r['entity_id'],'iso3':r.get('iso3'),'name_ja':r.get('name_ja'),'metrics_loaded':r.get('school_meals_metrics_loaded_count',0),'status':r.get('data_status')} for r in records if r.get('school_meals_metrics_loaded_count',0)<2]
    save_json(DATA/'market_base_school_meals_coverage_report_v103.json', coverage)
    save_json(DATA/'market_base_school_meals_source_manifest_v103.json', {'meta':{'version':VERSION,'created_at':CREATED_AT},'sources':manifest})
    print('[school] records:', len(records))
    print('[school] coverage:', coverage['metric_coverage'])

if __name__ == '__main__':
    main()
